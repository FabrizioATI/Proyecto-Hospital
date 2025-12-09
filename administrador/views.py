from django.shortcuts import render, redirect
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.db.models import Count
from database.models import Especialidad, Rol, Cita,WaitlistItem  

# ======= función utilitaria opcional (auto-ajustar columnas) =======
def autosize_columns(ws):
    for column_cells in ws.columns:
        length = 0
        col = column_cells[0].column  # número de columna
        for cell in column_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > length:
                    length = cell_len
            except:
                pass
        ws.column_dimensions[get_column_letter(col)].width = length + 2


# Index
def index(request):
    return render(request, "home/index.html")


# CRUD Mantenimiento
def mantenimiento_roles_especialidades(request):
    if request.method == "POST":
        accion = request.POST.get("accion")

        # ----------- ROLES -----------
        if accion == "agregar_rol":
            nombre = request.POST.get("nuevo_rol")
            if nombre:
                Rol.objects.create(nombre_rol=nombre)

        elif accion == "editar_rol":
            rol_id = request.POST.get("rol_id")
            nuevo_nombre = request.POST.get("editar_rol")
            if rol_id and nuevo_nombre:
                rol = Rol.objects.get(id=rol_id)
                rol.nombre_rol = nuevo_nombre
                rol.save()

        elif accion == "eliminar_rol":
            rol_id = request.POST.get("rol_id")
            if rol_id:
                Rol.objects.filter(id=rol_id).delete()

        # ----------- ESPECIALIDADES -----------
        elif accion == "agregar_especialidad":
            nombre = request.POST.get("nueva_especialidad")
            if nombre:
                Especialidad.objects.create(nombre=nombre)

        elif accion == "editar_especialidad":
            esp_id = request.POST.get("especialidad_id")
            nuevo_nombre = request.POST.get("editar_especialidad")
            if esp_id and nuevo_nombre:
                esp = Especialidad.objects.get(id=esp_id)
                esp.nombre = nuevo_nombre
                esp.save()

        elif accion == "eliminar_especialidad":
            esp_id = request.POST.get("especialidad_id")
            if esp_id:
                Especialidad.objects.filter(id=esp_id).delete()

        # Redirigir para evitar reenvío de formulario
        return redirect("mantenimiento_roles_especialidades")

    # GET: cargar listas
    roles = Rol.objects.all()
    especialidades = Especialidad.objects.all()

    return render(request, "mantenimiento/mantenimiento_roles_especialidades.html", {
        "roles": roles,
        "especialidades": especialidades,
    })


# =========================
# REPORTES EN EXCEL
# =========================
def vista_reportes(request):
    return render(request, "reportes/reportes.html")

def export_roles_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Roles"

    # Encabezados
    headers = ["ID", "Código Rol", "Nombre del Rol"]
    ws.append(headers)

    # Datos
    for rol in Rol.objects.all().order_by("id"):
        ws.append([
            rol.id,
            getattr(rol, "codigo_rol", ""),
            rol.nombre_rol,
        ])

    autosize_columns(ws)

    # Respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="roles.xlsx"'
    wb.save(response)
    return response

def export_especialidades_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Especialidades"

    # Encabezados
    headers = [
        "ID",
        "Nombre",
        "Capacidad por hora",
        "Requiere derivación",
    ]
    ws.append(headers)

    # Datos
    for esp in Especialidad.objects.all().order_by("id"):
        ws.append([
            esp.id,
            esp.nombre,
            esp.capacidad_por_hora,
            "Sí" if esp.requiere_derivacion else "No",
        ])

    autosize_columns(ws)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="especialidades.xlsx"'
    wb.save(response)
    return response

# 1) Resumen de citas por estado (EN_ESPERA, ATENDIDA, CANCELADA, NO_SHOW, etc.)
def export_citas_por_estado_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Citas por estado"

    headers = ["Estado", "Total de citas"]
    ws.append(headers)

    qs = (
        Cita.objects
        .values("estado")
        .annotate(total=Count("id"))
        .order_by("estado")
    )

    for row in qs:
        ws.append([
            row["estado"],
            row["total"],
        ])

    autosize_columns(ws)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="citas_por_estado.xlsx"'
    wb.save(response)
    return response


# 2) Citas por especialidad y tipo de cita (virtual/presencial)
def export_citas_por_especialidad_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Citas por especialidad"

    headers = ["Especialidad", "Tipo de cita", "Total de citas"]
    ws.append(headers)

    qs = (
        Cita.objects
        .select_related(
            "doctor_horario__doctor__especialidad"
        )
        .values(
            "doctor_horario__doctor__especialidad__nombre",
            "tipo_cita",
        )
        .annotate(total=Count("id"))
        .order_by("doctor_horario__doctor__especialidad__nombre", "tipo_cita")
    )

    for row in qs:
        ws.append([
            row["doctor_horario__doctor__especialidad__nombre"] or "Sin especialidad",
            row["tipo_cita"],
            row["total"],
        ])

    autosize_columns(ws)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="citas_por_especialidad.xlsx"'
    wb.save(response)
    return response


# 3) No-shows por doctor (quién tiene más pacientes que no asisten)
def export_noshow_por_doctor_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "No-show por doctor"

    headers = [
        "ID Doctor",
        "Nombre",
        "Apellido Paterno",
        "Apellido Materno",
        "Total NO_SHOW",
    ]
    ws.append(headers)

    qs = (
        Cita.objects
        .filter(estado="NO_SHOW")
        .values(
            "doctor__id",
            "doctor__nombre",
            "doctor__apellidoPaterno",
            "doctor__apellidoMaterno",
        )
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    for row in qs:
        ws.append([
            row["doctor__id"],
            row["doctor__nombre"],
            row["doctor__apellidoPaterno"],
            row["doctor__apellidoMaterno"],
            row["total"],
        ])

    autosize_columns(ws)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="noshow_por_doctor.xlsx"'
    wb.save(response)
    return response


# 4) Cancelaciones por paciente (pacientes que más cancelan)
def export_cancelaciones_por_paciente_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cancelaciones por paciente"

    headers = [
        "ID Paciente",
        "Nombre completo",
        "DNI",
        "Total CANCELADA",
    ]
    ws.append(headers)

    qs = (
        Cita.objects
        .filter(estado="CANCELADA")
        .values(
            "paciente__id",
            "paciente__nombre",
            "paciente__apellidoPaterno",
            "paciente__apellidoMaterno",
            "paciente__dni",
        )
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    for row in qs:
        nombre_completo = f"{row['paciente__nombre']} {row['paciente__apellidoPaterno']} {row['paciente__apellidoMaterno']}"
        ws.append([
            row["paciente__id"],
            nombre_completo,
            row["paciente__dni"],
            row["total"],
        ])

    autosize_columns(ws)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="cancelaciones_por_paciente.xlsx"'
    wb.save(response)
    return response


# 5) Resumen de lista de espera (pendiente, notificado, aceptado, expirado, rechazado)
def export_waitlist_resumen_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Waitlist resumen"

    headers = ["Estado en lista de espera", "Total"]
    ws.append(headers)

    qs = (
        WaitlistItem.objects
        .values("estado")
        .annotate(total=Count("id"))
        .order_by("estado")
    )

    for row in qs:
        ws.append([
            row["estado"],
            row["total"],
        ])

    autosize_columns(ws)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="waitlist_resumen.xlsx"'
    wb.save(response)
    return response
# ========= EJEMPLO EXTRA: REPORTE DE CITAS =========
# Puedes copiar este patrón para WaitlistItem, SMSNotification, etc.

def export_citas_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Citas"

    headers = [
        "ID",
        "Paciente",
        "DNI Paciente",
        "Doctor",
        "Especialidad",
        "Fecha",
        "Hora inicio",
        "Hora fin",
        "Clasificación",
        "Prioridad",
        "Estado",
        "Tipo cita",
        "Motivo",
        "EHR ID",
        "Fecha creación",
    ]
    ws.append(headers)

    # Usamos select_related para evitar N+1
    from database.models import DoctorHorario  # por si quieres usarlo directo

    citas = (
        Cita.objects
        .select_related(
            "paciente",
            "doctor",
            "doctor_horario",
            "doctor_horario__doctor",
            "doctor_horario__doctor__entidad",
            "doctor_horario__doctor__especialidad",
            "doctor_horario__horario",
        )
        .all()
    )

    for c in citas:
        horario = c.doctor_horario.horario
        doctor_detalle = c.doctor_horario.doctor
        especialidad = doctor_detalle.especialidad.nombre if doctor_detalle and doctor_detalle.especialidad else ""

        ws.append([
            c.id,
            c.paciente.nombre_completo(),
            c.paciente.dni,
            c.doctor.nombre_completo(),
            especialidad,
            horario.fecha if horario else "",
            horario.hora_inicio if horario else "",
            horario.hora_fin if horario else "",
            c.clasificacion,
            c.prioridad,
            c.estado,
            c.tipo_cita,
            c.motivo,
            c.ehr_id,
            c.fecha_creacion,
        ])

    autosize_columns(ws)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="citas.xlsx"'
    wb.save(response)
    return response
